"""Sample data fixtures for testing."""

import pandas as pd
from pathlib import Path
from unittest.mock import Mock

from cfobot.data_loader import FinancialData


def create_sample_balance_data():
    """Create sample balance sheet data."""
    return pd.DataFrame({
        'Nivel': ['Clase', 'Clase', 'Clase', 'Grupo', 'Grupo', 'Grupo', 'Grupo', 'Grupo'],
        'Código cuenta contable': ['1', '2', '3', '11', '12', '13', '14', '21'],
        'Nombre cuenta contable': [
            'ACTIVO', 'PASIVO', 'PATRIMONIO', 
            'ACTIVO CORRIENTE', 'INVERSIONES', 'CUENTAS POR COBRAR', 'INVENTARIOS', 'PASIVO CORRIENTE'
        ],
        'Saldo inicial': [0, 0, 0, 0, 0, 0, 0, 0],
        'Movimiento débito': [0, 0, 0, 0, 0, 0, 0, 0],
        'Movimiento crédito': [0, 0, 0, 0, 0, 0, 0, 0],
        'Saldo final': [500000000, 200000000, 300000000, 150000000, 50000000, 80000000, 120000000, 200000000]
    })


def create_sample_eri_data():
    """Create sample ERI (Income and Expense Report) data."""
    return pd.DataFrame({
        'Codigo': ['510101', '510201', '530101', '610101', '720101', '510301', '510401'],
        'Nombre': [
            'SUELDOS ADMINISTRATIVOS', 'CESANTIAS', 'INTERESES', 
            'COSTO VENTAS', 'COSTO PRODUCCION', 'DEPRECIACION', 'OTROS GASTOS'
        ],
        'Display Name': [
            'SUELDOS ADMINISTRATIVOS', 'CESANTIAS', 'INTERESES', 
            'COSTO VENTAS', 'COSTO PRODUCCION', 'DEPRECIACION', 'OTROS GASTOS'
        ],
        'ENERO DE 2025': [50000000, 10000000, 5000000, 30000000, 40000000, 8000000, 15000000],
        'FEBRERO DE 2025': [52000000, 10500000, 4800000, 32000000, 42000000, 8200000, 16000000],
        'MARZO DE 2025': [48000000, 9800000, 5200000, 28000000, 38000000, 7800000, 14000000]
    })


def create_sample_income_statement_data():
    """Create sample income statement data."""
    return pd.DataFrame({
        'Descripcion': [
            'INGRESOS ORDINARIOS', 
            'COSTO DE VENTA', 
            'RESULTADO DEL EJERCICIO'
        ],
        'Total ENERO': [120000000, 30000000, 15000000],
        'Total FEBRERO': [125000000, 32000000, 18000000],
        'Total MARZO': [110000000, 28000000, 12000000]
    })


def create_sample_caratula_data():
    """Create sample caratula (cover page) data."""
    return pd.DataFrame({
        'Column_0': ['Diferencia', 'Otro dato'],
        'Column_1': [5000000, 0]
    })


def create_sample_financial_data():
    """Create complete sample financial data."""
    # Mock workbook
    mock_workbook = Mock()
    mock_workbook.sheet_names = [
        'BALANCE ENERO', 'BALANCE FEBRERO', 'BALANCE MARZO', 
        'INFORME-ERI', 'ESTADO RESULTADO', 'CARATULA'
    ]
    
    return FinancialData(
        current_month='MARZO',
        current_month_col='MARZO DE 2025',
        resultado_current_col='Total MARZO',
        balance=create_sample_balance_data(),
        eri=create_sample_eri_data(),
        resultado=create_sample_income_statement_data(),
        caratula=create_sample_caratula_data(),
        months=['ENERO DE 2025', 'FEBRERO DE 2025', 'MARZO DE 2025'],
        workbook=mock_workbook
    )


def create_sample_excel_file(tmp_path: Path) -> Path:
    """Create a sample Excel file for testing."""
    import openpyxl
    from openpyxl import Workbook
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    # Balance sheets for prior and current months
    balance_data = create_sample_balance_data()
    for sheet_name in ["BALANCE FEBRERO", "BALANCE MARZO"]:
        balance_ws = wb.create_sheet(sheet_name)
        for r, row in enumerate(balance_data.values, start=5):  # Start at row 5 (skiprows=4)
            for c, value in enumerate(row, start=1):
                balance_ws.cell(row=r, column=c, value=value)
    
    # ERI sheet
    eri_ws = wb.create_sheet("INFORME-ERI")
    eri_data = create_sample_eri_data()
    for r, row in enumerate(eri_data.values, start=2):  # Start at row 2 (skiprows=1)
        for c, value in enumerate(row, start=1):
            eri_ws.cell(row=r, column=c, value=value)
    
    # Income statement sheet
    income_ws = wb.create_sheet("ESTADO RESULTADO")
    income_data = create_sample_income_statement_data()
    # Add headers for multi-level columns
    income_ws.cell(row=1, column=1, value="Descripcion")
    income_ws.cell(row=1, column=2, value="ENERO")
    income_ws.cell(row=1, column=3, value="FEBRERO")
    income_ws.cell(row=1, column=4, value="MARZO")
    income_ws.cell(row=2, column=1, value="Descripcion")
    income_ws.cell(row=2, column=2, value="Total ENERO")
    income_ws.cell(row=2, column=3, value="Total FEBRERO")
    income_ws.cell(row=2, column=4, value="Total MARZO")
    
    for r, row in enumerate(income_data.values, start=3):  # Start at row 3 (skiprows=2)
        for c, value in enumerate(row, start=1):
            income_ws.cell(row=r, column=c, value=value)
    
    # Caratula sheet
    caratula_ws = wb.create_sheet("CARATULA")
    caratula_data = create_sample_caratula_data()
    for r, row in enumerate(caratula_data.values, start=6):  # Start at row 6 (skiprows=5)
        for c, value in enumerate(row, start=1):
            caratula_ws.cell(row=r, column=c, value=value)
    
    # Save file
    file_path = tmp_path / "INFORME DE MARZO APRU- 2025 .xlsx"
    wb.save(file_path)
    return file_path


def create_sample_config():
    """Create sample configuration."""
    from cfobot.config import AppConfig, BudgetConfig
    
    return AppConfig(
        budgets=BudgetConfig(
            ingresos_mensual=100_000_000,
            gastos_mensual=125_000_000
        )
    )


def create_sample_email_config():
    """Create sample email configuration."""
    from cfobot.config import EmailConfig
    
    return EmailConfig(
        smtp_server="smtp.gmail.com",
        smtp_port=587,
        sender_email="test@example.com",
        sender_password="test_password",
        recipient_emails=["recipient1@example.com", "recipient2@example.com"]
    )


# Test data for edge cases
def create_empty_financial_data():
    """Create empty financial data for testing edge cases."""
    mock_workbook = Mock()
    mock_workbook.sheet_names = []
    
    return FinancialData(
        current_month='MARZO',
        current_month_col='MARZO DE 2025',
        resultado_current_col='Total MARZO',
        balance=pd.DataFrame(),
        eri=pd.DataFrame(),
        resultado=pd.DataFrame(),
        caratula=pd.DataFrame(),
        months=[],
        workbook=mock_workbook
    )


def create_malformed_financial_data():
    """Create malformed financial data for testing error handling."""
    mock_workbook = Mock()
    mock_workbook.sheet_names = ['BALANCE MARZO', 'INFORME-ERI', 'ESTADO RESULTADO', 'CARATULA']
    
    # Create data with missing required columns
    malformed_balance = pd.DataFrame({
        'Wrong_Column': [1, 2, 3],
        'Another_Wrong_Column': ['A', 'B', 'C']
    })
    
    malformed_eri = pd.DataFrame({
        'Wrong_Code': ['X', 'Y', 'Z'],
        'Wrong_Name': ['Item1', 'Item2', 'Item3']
    })
    
    malformed_resultado = pd.DataFrame({
        'Wrong_Description': ['Item1', 'Item2'],
        'Wrong_Total': [100, 200]
    })
    
    return FinancialData(
        current_month='MARZO',
        current_month_col='MARZO DE 2025',
        resultado_current_col='Total MARZO',
        balance=malformed_balance,
        eri=malformed_eri,
        resultado=malformed_resultado,
        caratula=pd.DataFrame(),
        months=[],
        workbook=mock_workbook
    )
