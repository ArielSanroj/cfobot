"""Integration tests for the complete CFO Bot pipeline."""

import pytest
import pandas as pd
from pathlib import Path
from unittest.mock import Mock, patch

from cfobot.cli import run_pipeline
from cfobot.config import AppConfig, BudgetConfig, EmailConfig
from tests.fixtures.sample_data import (
    create_sample_financial_data,
    create_sample_excel_file,
    create_sample_config,
    create_sample_email_config
)


class TestCompletePipeline:
    """Test the complete CFO Bot pipeline."""
    
    @patch('cfobot.cli.find_latest_report')
    @patch('cfobot.cli.load_financial_data')
    @patch('cfobot.cli.consolidate_balance')
    @patch('cfobot.cli.compute_budget_execution')
    @patch('cfobot.cli.compute_kpis')
    @patch('cfobot.cli.save_consolidated_balance')
    @patch('cfobot.cli.save_budget_execution')
    @patch('cfobot.cli.save_kpis')
    @patch('cfobot.cli.generate_all_figures')
    @patch('cfobot.cli.extract_caratula_difference')
    @patch('cfobot.cli.build_board_report')
    def test_complete_pipeline_success(
        self,
        mock_build_board_report,
        mock_extract_caratula,
        mock_generate_figures,
        mock_save_kpis,
        mock_save_budget,
        mock_save_consolidated,
        mock_compute_kpis,
        mock_compute_budget,
        mock_consolidate_balance,
        mock_load_data,
        mock_find_report
    ):
        """Test successful complete pipeline execution."""
        # Setup mocks
        mock_find_report.return_value = Path("test_report.xlsx")
        mock_load_data.return_value = create_sample_financial_data()
        mock_consolidate_balance.return_value = pd.DataFrame()
        mock_compute_budget.return_value = Mock()
        mock_compute_kpis.return_value = Mock()
        mock_save_consolidated.return_value = Path("consolidated.xlsx")
        mock_save_budget.return_value = Path("budget.xlsx")
        mock_save_kpis.return_value = Path("kpis.xlsx")
        mock_generate_figures.return_value = [Path("chart1.png"), Path("chart2.png")]
        mock_extract_caratula.return_value = 5000000.0
        mock_build_board_report.return_value = Path("board_report.docx")
        
        # Create config
        config = create_sample_config()
        
        # Run pipeline
        run_pipeline(config=config, send_email=False, skip_visuals=False)
        
        # Verify all functions were called
        mock_find_report.assert_called_once()
        mock_load_data.assert_called_once()
        mock_consolidate_balance.assert_called_once()
        mock_compute_budget.assert_called_once()
        mock_compute_kpis.assert_called_once()
        mock_save_consolidated.assert_called_once()
        mock_save_budget.assert_called_once()
        mock_save_kpis.assert_called_once()
        mock_generate_figures.assert_called_once()
        mock_extract_caratula.assert_called_once()
        mock_build_board_report.assert_called_once()
    
    @patch('cfobot.cli.find_latest_report')
    @patch('cfobot.cli.load_financial_data')
    @patch('cfobot.cli.consolidate_balance')
    @patch('cfobot.cli.compute_budget_execution')
    @patch('cfobot.cli.compute_kpis')
    @patch('cfobot.cli.save_consolidated_balance')
    @patch('cfobot.cli.save_budget_execution')
    @patch('cfobot.cli.save_kpis')
    @patch('cfobot.cli.extract_caratula_difference')
    @patch('cfobot.cli.build_board_report')
    @patch('cfobot.cli.send_reports')
    def test_complete_pipeline_with_email(
        self,
        mock_send_reports,
        mock_build_board_report,
        mock_extract_caratula,
        mock_save_kpis,
        mock_save_budget,
        mock_save_consolidated,
        mock_compute_kpis,
        mock_compute_budget,
        mock_consolidate_balance,
        mock_load_data,
        mock_find_report
    ):
        """Test complete pipeline with email sending."""
        # Setup mocks
        mock_find_report.return_value = Path("test_report.xlsx")
        mock_load_data.return_value = create_sample_financial_data()
        mock_consolidate_balance.return_value = pd.DataFrame()
        mock_compute_budget.return_value = Mock()
        mock_compute_kpis.return_value = Mock()
        mock_save_consolidated.return_value = Path("consolidated.xlsx")
        mock_save_budget.return_value = Path("budget.xlsx")
        mock_save_kpis.return_value = Path("kpis.xlsx")
        mock_extract_caratula.return_value = 5000000.0
        mock_build_board_report.return_value = Path("board_report.docx")
        
        # Create config with email
        config = create_sample_config()
        config.email = create_sample_email_config()
        
        # Run pipeline with email
        run_pipeline(config=config, send_email=True, skip_visuals=False)
        
        # Verify email was sent
        mock_send_reports.assert_called_once()
        call_args = mock_send_reports.call_args
        assert call_args[0][0] == config.email  # email_config
        assert "Reporte CFO Automatizado" in call_args[0][1]  # subject
        assert "html" in call_args[0][2]  # html_body
    
    @patch('cfobot.cli.find_latest_report')
    @patch('cfobot.cli.load_financial_data')
    @patch('cfobot.cli.consolidate_balance')
    @patch('cfobot.cli.compute_budget_execution')
    @patch('cfobot.cli.compute_kpis')
    @patch('cfobot.cli.save_consolidated_balance')
    @patch('cfobot.cli.save_budget_execution')
    @patch('cfobot.cli.save_kpis')
    @patch('cfobot.cli.extract_caratula_difference')
    @patch('cfobot.cli.build_board_report')
    def test_complete_pipeline_skip_visuals(
        self,
        mock_build_board_report,
        mock_extract_caratula,
        mock_save_kpis,
        mock_save_budget,
        mock_save_consolidated,
        mock_compute_kpis,
        mock_compute_budget,
        mock_consolidate_balance,
        mock_load_data,
        mock_find_report
    ):
        """Test complete pipeline with visuals skipped."""
        # Setup mocks
        mock_find_report.return_value = Path("test_report.xlsx")
        mock_load_data.return_value = create_sample_financial_data()
        mock_consolidate_balance.return_value = pd.DataFrame()
        mock_compute_budget.return_value = Mock()
        mock_compute_kpis.return_value = Mock()
        mock_save_consolidated.return_value = Path("consolidated.xlsx")
        mock_save_budget.return_value = Path("budget.xlsx")
        mock_save_kpis.return_value = Path("kpis.xlsx")
        mock_extract_caratula.return_value = 5000000.0
        mock_build_board_report.return_value = Path("board_report.docx")
        
        # Create config
        config = create_sample_config()
        
        # Run pipeline with visuals skipped
        run_pipeline(config=config, send_email=False, skip_visuals=True)
        
        # Verify all functions were called except figure generation
        mock_find_report.assert_called_once()
        mock_load_data.assert_called_once()
        mock_consolidate_balance.assert_called_once()
        mock_compute_budget.assert_called_once()
        mock_compute_kpis.assert_called_once()
        mock_save_consolidated.assert_called_once()
        mock_save_budget.assert_called_once()
        mock_save_kpis.assert_called_once()
        mock_extract_caratula.assert_called_once()
        mock_build_board_report.assert_called_once()
    
    @patch('cfobot.cli.find_latest_report')
    def test_pipeline_file_not_found(self, mock_find_report):
        """Test pipeline when report file is not found."""
        from cfobot.cli import run_pipeline
        
        # Setup mock to raise FileNotFoundError
        mock_find_report.side_effect = FileNotFoundError("No file found")
        
        config = create_sample_config()
        
        # Should raise FileNotFoundError
        with pytest.raises(FileNotFoundError):
            run_pipeline(config=config, send_email=False, skip_visuals=False)
    
    @patch('cfobot.cli.find_latest_report')
    @patch('cfobot.cli.load_financial_data')
    def test_pipeline_invalid_data(self, mock_load_data, mock_find_report):
        """Test pipeline with invalid data."""
        from cfobot.cli import run_pipeline
        
        # Setup mocks
        mock_find_report.return_value = Path("test_report.xlsx")
        mock_load_data.side_effect = ValueError("Invalid data format")
        
        config = create_sample_config()
        
        # Should raise ValueError
        with pytest.raises(ValueError):
            run_pipeline(config=config, send_email=False, skip_visuals=False)
    
    @patch('cfobot.cli.find_latest_report')
    @patch('cfobot.cli.load_financial_data')
    @patch('cfobot.cli.consolidate_balance')
    @patch('cfobot.cli.compute_budget_execution')
    @patch('cfobot.cli.compute_kpis')
    @patch('cfobot.cli.save_consolidated_balance')
    @patch('cfobot.cli.save_budget_execution')
    @patch('cfobot.cli.save_kpis')
    @patch('cfobot.cli.extract_caratula_difference')
    @patch('cfobot.cli.build_board_report')
    def test_pipeline_with_email_config_missing(
        self,
        mock_build_board_report,
        mock_extract_caratula,
        mock_save_kpis,
        mock_save_budget,
        mock_save_consolidated,
        mock_compute_kpis,
        mock_compute_budget,
        mock_consolidate_balance,
        mock_load_data,
        mock_find_report
    ):
        """Test pipeline with email requested but no email config."""
        # Setup mocks
        mock_find_report.return_value = Path("test_report.xlsx")
        mock_load_data.return_value = create_sample_financial_data()
        mock_consolidate_balance.return_value = pd.DataFrame()
        mock_compute_budget.return_value = Mock()
        mock_compute_kpis.return_value = Mock()
        mock_save_consolidated.return_value = Path("consolidated.xlsx")
        mock_save_budget.return_value = Path("budget.xlsx")
        mock_save_kpis.return_value = Path("kpis.xlsx")
        mock_extract_caratula.return_value = 5000000.0
        mock_build_board_report.return_value = Path("board_report.docx")
        
        # Create config without email
        config = create_sample_config()
        config.email = None
        
        # Run pipeline with email requested but no config
        run_pipeline(config=config, send_email=True, skip_visuals=False)
        
        # Should complete without sending email
        mock_find_report.assert_called_once()
        mock_load_data.assert_called_once()
        mock_consolidate_balance.assert_called_once()
        mock_compute_budget.assert_called_once()
        mock_compute_kpis.assert_called_once()
        mock_save_consolidated.assert_called_once()
        mock_save_budget.assert_called_once()
        mock_save_kpis.assert_called_once()
        mock_extract_caratula.assert_called_once()
        mock_build_board_report.assert_called_once()


class TestExcelFileProcessing:
    """Test Excel file processing integration."""
    
    def test_excel_file_creation_and_processing(self, tmp_path):
        """Test creating and processing a sample Excel file."""
        # Create sample Excel file
        excel_file = create_sample_excel_file(tmp_path)
        assert excel_file.exists()
        
        # Test that the file can be read
        import pandas as pd
        workbook = pd.ExcelFile(excel_file)
        
        # Verify all required sheets exist
        required_sheets = ['BALANCE MARZO', 'INFORME-ERI', 'ESTADO RESULTADO', 'CARATULA']
        for sheet in required_sheets:
            assert sheet in workbook.sheet_names
        
        # Test reading each sheet
        balance_df = pd.read_excel(workbook, sheet_name='BALANCE MARZO', skiprows=4)
        eri_df = pd.read_excel(workbook, sheet_name='INFORME-ERI', skiprows=1)
        resultado_df = pd.read_excel(workbook, sheet_name='ESTADO RESULTADO', skiprows=2, header=[0, 1])
        caratula_df = pd.read_excel(workbook, sheet_name='CARATULA', skiprows=5)
        
        # Verify data structure
        assert not balance_df.empty
        assert not eri_df.empty
        assert not resultado_df.empty
        assert not caratula_df.empty
    
    def test_excel_file_with_missing_sheets(self, tmp_path):
        """Test Excel file with missing required sheets."""
        import openpyxl
        from openpyxl import Workbook
        
        # Create workbook with only some sheets
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("BALANCE MARZO")
        wb.create_sheet("INFORME-ERI")
        # Missing "ESTADO RESULTADO" and "CARATULA"
        
        file_path = tmp_path / "incomplete_report.xlsx"
        wb.save(file_path)
        
        # Test loading should fail
        from cfobot.data_loader import load_financial_data
        from cfobot.config import AppConfig
        
        config = AppConfig()
        
        with pytest.raises(ValueError, match="Required sheets missing"):
            load_financial_data(file_path, config, Mock())


class TestDataValidation:
    """Test data validation integration."""
    
    def test_balance_equation_validation(self):
        """Test balance equation validation with real data."""
        from cfobot.validators import validate_balance_equation
        
        # Create balanced data
        balanced_df = pd.DataFrame({
            'Nivel': ['Clase', 'Clase', 'Clase'],
            'Código cuenta contable': ['1', '2', '3'],
            'Saldo final': [1000, 600, 400]  # 1000 = 600 + 400
        })
        assert validate_balance_equation(balanced_df) is True
        
        # Create unbalanced data
        unbalanced_df = pd.DataFrame({
            'Nivel': ['Clase', 'Clase', 'Clase'],
            'Código cuenta contable': ['1', '2', '3'],
            'Saldo final': [1000, 600, 300]  # 1000 != 600 + 300
        })
        assert validate_balance_equation(unbalanced_df) is False
    
    def test_outlier_detection(self):
        """Test outlier detection with real data."""
        from cfobot.validators import detect_outliers
        
        # Normal data
        normal_series = pd.Series([100, 105, 98, 102, 103])
        outliers = detect_outliers(normal_series)
        assert outliers.sum() == 0
        
        # Data with outlier
        outlier_series = pd.Series([100, 105, 98, 1000, 102])
        outliers = detect_outliers(outlier_series)
        assert outliers.sum() == 1
        assert outliers.iloc[3] is True
